from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "japan-itinerary.xlsx"
FONT = "Arial"

INK       = "1A1C2A"
INK_SOFT  = "3A3E52"
VERM      = "B2431C"
GOLD      = "A8864A"
PAPER     = "F2ECDE"
PAPER_WARM= "ECE3D0"
RULE      = "C7BFAE"
WHITE     = "FFFFFF"

thin = Side(style="thin", color=RULE)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = Font(name=FONT, bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", start_color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border_all

def style_body(cell, alt=False, bold=False, color=None):
    cell.font = Font(name=FONT, size=10, bold=bold, color=color or INK)
    cell.fill = PatternFill("solid", start_color=(PAPER_WARM if alt else PAPER))
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_all

def style_title(cell):
    cell.font = Font(name=FONT, size=16, bold=True, color=VERM)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_sub(cell):
    cell.font = Font(name=FONT, size=10, italic=True, color=INK_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center")

wb = Workbook()

# ============================================================
# SHEET 1 — ITINERARY
# ============================================================
s1 = wb.active
s1.title = "Itinerary"

s1["A1"] = "Japan, softly"
style_title(s1["A1"])
s1.merge_cells("A1:J1")

s1["A2"] = "An itinerary for Anushka & Robin — May 12 to 22, 2026"
style_sub(s1["A2"])
s1.merge_cells("A2:J2")

headers = ["Day", "Date", "City", "Theme", "Morning", "Afternoon", "Evening", "Food picks", "Reservations", "Notes"]
for col, h in enumerate(headers, 1):
    c = s1.cell(row=4, column=col, value=h)
    style_header(c)

rows = [
    ["1 · Tue",  "May 12",  "Tokyo",
     "A soft landing",
     "Narita → hotel (Limousine Bus or private car, ~90 min)",
     "Check in at Prince Gallery Tokyo Kioicho; unpack slowly",
     "Hie Shrine at dusk · tonkatsu at Butagumi Shokudo",
     "Tonkatsu (rosu + hire)",
     "—",
     "Activate Suica on Apple Wallet at the airport"],

    ["2 · Wed",  "May 13",  "Tokyo",
     "Classic Tokyo, unrushed",
     "Tsukiji Outer Market breakfast · Meiji Jingu",
     "Takeshita/Cat Street · Afuri ramen · Omotesando stroll",
     "Shibuya Sky sunset · cross Shibuya · Omoide Yokocho (Kabuto yakitori)",
     "Tamagoyaki · fruit sando · yuzu-shio ramen · grilled eel",
     "Shibuya Sky sunset slot",
     "Rush hour 07:30–09:30 — avoid Yamanote"],

    ["3 · Thu",  "May 14",  "Tokyo",
     "A day at sea",
     "Taxi to Maihama · DisneySea rope drop · Fantasy Springs (Peter Pan, Frozen)",
     "Originals: Sinbad, Journey, Tower of Terror · lunch Gyoza Dog / Tipo Torta",
     "Fantastic Flight · 20,000 Leagues · Believe! Sea of Dreams nighttime show",
     "Gyoza Dog · Tipo Torta · Curry popcorn",
     "DisneySea 1-Day Passport (book 60+ days out)",
     "Premier Access on app at 07:00 day-of"],

    ["4 · Fri",  "May 15",  "Tokyo",
     "Old Tokyo & Michelin ★★",
     "Cooking class (Buddha Bellies or Cooking Sun) · dashi + tamagoyaki",
     "Yanaka Ginza · Nezu Shrine azalea tunnels",
     "Den — Michelin ★★ · slow walk home",
     "Menchi-katsu · taiyaki · iced matcha · Den tasting menu",
     "Den (60+ days out · Pocket Concierge or TableAll)",
     "Smart casual · 2.5–3 hours"],

    ["5 · Sat",  "May 16",  "Tokyo",
     "Bright pleasures",
     "Pokémon Center + Nintendo Tokyo (Shibuya Parco)",
     "Kirby Café Tokyo (Skytree Town) · café hour at Azabudai Hills",
     "teamLab Borderless (last slot) · Uobei sushi or Sarashina Horii soba",
     "Kirby curry · Maxim Tomato pasta · conveyor sushi",
     "Kirby Café (1 month to the minute) · teamLab",
     "Nintendo Tokyo ships overseas for large items"],

    ["6 · Sun",  "May 17",  "Tokyo → Osaka",
     "Shinkansen south",
     "Pack, Japanese breakfast · Yamato forwarding the night before",
     "Taxi Shinagawa · 12:30 Nozomi to Shin-Osaka (2h25m) · Makunouchi bento",
     "Arrive Osaka Station Hotel · Midosuji to Dotonbori street crawl",
     "Takoyaki (Kukuru) · okonomiyaki (Mizuno) · kushikatsu (Daruma) · 551 pork bun · Rikuro cheesecake",
     "SmartEX app · reserve E-window for Fuji",
     "Luggage forwarding ¥2,200/bag, arrives before you"],

    ["7 · Mon",  "May 18",  "Kyoto",
     "Kyoto, the first time · Michelin ★★★",
     "07:00 JR Special Rapid to Kyoto · Fushimi Inari 10,000 gates",
     "Inari-zushi at Neneya · Kiyomizu-dera · Sannenzaka/Ninenzaka · Nishiki Market lunch",
     "Gion stroll (Hanami-koji, Shirakawa) · Kikunoi Honten ★★★ kaiseki",
     "Inari-zushi · dashimaki tamago · yuba · warabi mochi · Kikunoi kaiseki",
     "Kikunoi (60+ days · concierge)",
     "Please don't photograph geiko on the way to work"],

    ["8 · Tue",  "May 19",  "Kyoto",
     "Bamboo & a golden pavilion",
     "JR Sagano to Saga-Arashiyama · Bamboo Grove · Tenryū-ji garden",
     "Shigetsu temple lunch (shōjin ryōri) · taxi to Kinkaku-ji",
     "Matcha ceremony (Camellia or En) · Philosopher's Path to Ginkaku-ji",
     "Shōjin ryōri · matcha soft serve · soba at Honke Owariya",
     "Shigetsu lunch · tea ceremony (1 week out)",
     "Alt: Sagano Scenic Train instead of Kinkaku-ji"],

    ["9 · Wed",  "May 20",  "Nara & Osaka",
     "Bowing deer, afternoon off",
     "JR Yamatoji Rapid to Nara (45 min) · Nara Park deer · Tōdai-ji Great Buddha",
     "Nakatanidō mochi · kakinoha-zushi at Hiraso · back to Osaka",
     "Yours: Shinsaibashi / Grand Front / hotel lounge · sushi at Harasho (Fukushima)",
     "Shika-senbei crackers · mochi · persimmon-leaf sushi · omakase",
     "Sushi Harasho (2 weeks out)",
     "Alt: Himeji Castle (40 min Shinkansen)"],

    ["10 · Thu", "May 21", "Osaka",
     "Mario's own park",
     "JR Yumesaki to Universal City (12 min) · Super Nintendo World rope drop · Mario Kart",
     "Yoshi's Adventure · Kinopio's Café (Piranha Plant pizza)",
     "Pick 2 from USJ · home · final Osaka okonomiyaki (Chibo or Mizuno) · melon pan ice cream",
     "Piranha Plant pizza · Yoshi calzone · okonomiyaki · melon pan sandwich",
     "USJ ticket + Express Pass 4 Nintendo",
     "Buy Power-Up Band at entrance (¥4,200)"],

    ["11 · Fri", "May 22", "Osaka → Home",
     "Home, via Tokyo",
     "Breakfast to go · JR to Shin-Osaka (4 min)",
     "08:00 Nozomi to Shinagawa (arrive 10:30) · Narita Express (N'EX, 75 min)",
     "Check in Narita (3 hrs before flight) · ekiben & Kit Kat flight pack",
     "Onigiri · ekiben · green tea Kit Kat",
     "Shinkansen (SmartEX, 2–3 days out) · N'EX reservation",
     "Allow 4.5 hrs hotel → airport to be safe"],
]

for r_idx, row in enumerate(rows, start=5):
    alt = r_idx % 2 == 1
    for c_idx, val in enumerate(row, start=1):
        cell = s1.cell(row=r_idx, column=c_idx, value=val)
        style_body(cell, alt=alt, bold=(c_idx == 1))
        if c_idx == 1:
            cell.font = Font(name=FONT, size=10, bold=True, color=VERM)

widths = [10, 10, 16, 26, 34, 34, 40, 32, 28, 34]
for i, w in enumerate(widths, 1):
    s1.column_dimensions[get_column_letter(i)].width = w

# Tall rows for the multi-line text
s1.row_dimensions[1].height = 28
for r in range(5, 5 + len(rows)):
    s1.row_dimensions[r].height = 78

s1.freeze_panes = "A5"
s1.print_title_rows = "1:4"
s1.page_setup.orientation = s1.ORIENTATION_LANDSCAPE
s1.page_setup.fitToWidth = 1
s1.page_setup.fitToHeight = 0
s1.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================
# SHEET 2 — RESERVATIONS
# ============================================================
s2 = wb.create_sheet("Reservations")

s2["A1"] = "Reservations to make now"
style_title(s2["A1"])
s2.merge_cells("A1:F1")

s2["A2"] = "The only things in the trip that are genuinely hard to book — handle these first."
style_sub(s2["A2"])
s2.merge_cells("A2:F2")

r_headers = ["Book by", "What", "Day", "Details", "Lead time", "Status"]
for col, h in enumerate(r_headers, 1):
    c = s2.cell(row=4, column=col, value=h)
    style_header(c)

res_rows = [
    ["Mar 15", "Den — Michelin ★★ kaiseki", "Fri May 15", "Pocket Concierge or TableAll; ¥38,000 pp; 2.5–3 hrs", "60 days",   ""],
    ["Mar 18", "Kikunoi Honten — Michelin ★★★", "Mon May 18", "Hotel concierge or email direct; ¥44,000 pp; 2.5 hrs", "60+ days", ""],
    ["Mar 20", "Tokyo DisneySea — 1-Day Passport x2", "Thu May 14", "tokyodisneyresort.jp; ¥10,900 pp", "60 days", ""],
    ["Apr 14", "Kirby Café Tokyo", "Sat May 16", "Opens exactly 1 month out at 18:00 JST; fills in minutes", "1 month", ""],
    ["Apr 14", "USJ park ticket + Super Nintendo timed entry", "Thu May 21", "USJ app; buy Express Pass 4: Nintendo + Hollywood Dream", "1–2 months", ""],
    ["Apr 20", "Private cooking class, Tokyo", "Fri May 15 AM", "Buddha Bellies or Cooking Sun", "1 month", ""],
    ["Apr 25", "teamLab Borderless (Azabudai)", "Sat May 16 PM", "teamlab.art/e/borderless-azabudai; last evening slot", "2 weeks", ""],
    ["Apr 28", "Shibuya Sky sunset slot", "Wed May 13 PM", "shibuya-scramble-square.com", "2 weeks", ""],
    ["May 5",  "Matcha tea ceremony (Philosopher's Path)", "Tue May 19 PM", "Camellia or En Tea House; ¥5,500 pp", "1 week", ""],
    ["May 10", "Sushi Harasho (Fukushima, Osaka)", "Wed May 20 PM", "Through hotel; ¥15,000 omakase, 10 seats", "2 weeks", ""],
    ["May 10", "Shinkansen — Nozomi Tokyo→Osaka", "Sun May 17",  "SmartEX app; reserve E-window for Fuji",  "2–3 days",  ""],
    ["May 20", "Shinkansen — Nozomi Osaka→Tokyo + Narita Express", "Fri May 22", "SmartEX for Nozomi; N'EX at JR East counter", "1–2 days", ""],
    ["Any",    "Yamato luggage forwarding, Tokyo → Osaka", "Sat May 16", "Prince Gallery concierge by 14:00 · ¥2,200 per bag", "Day before", ""],
]

for r_idx, row in enumerate(res_rows, start=5):
    alt = r_idx % 2 == 1
    for c_idx, val in enumerate(row, start=1):
        cell = s2.cell(row=r_idx, column=c_idx, value=val)
        style_body(cell, alt=alt, bold=(c_idx == 1))
        if c_idx == 1:
            cell.font = Font(name=FONT, size=10, bold=True, color=VERM)

r_widths = [12, 36, 14, 52, 14, 18]
for i, w in enumerate(r_widths, 1):
    s2.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 5 + len(res_rows)):
    s2.row_dimensions[r].height = 30

s2.freeze_panes = "A5"
s2.page_setup.orientation = s2.ORIENTATION_LANDSCAPE
s2.page_setup.fitToWidth = 1
s2.page_setup.fitToHeight = 0
s2.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================
# SHEET 3 — TRANSIT & PAYMENTS
# ============================================================
s3 = wb.create_sheet("Transit & Essentials")

s3["A1"] = "Moving through Japan"
style_title(s3["A1"])
s3.merge_cells("A1:E1")

s3["A2"] = "Payments, public transit, and the Shinkansen at a glance."
style_sub(s3["A2"])
s3.merge_cells("A2:E2")

# Transit table
s3["A4"] = "Intercity & day-trip trains"
s3["A4"].font = Font(name=FONT, size=12, bold=True, color=INK)
s3.merge_cells("A4:E4")

t_headers = ["Route", "Train", "Time", "Cost (one way)", "Notes"]
for col, h in enumerate(t_headers, 1):
    c = s3.cell(row=5, column=col, value=h)
    style_header(c)

transit_rows = [
    ["Tokyo ↔ Osaka",   "Nozomi (Shinkansen)",     "2h 25m",   "¥14,720", "Board at Shinagawa (easier) or Tokyo Station. E-window for Fuji."],
    ["Osaka ↔ Kyoto",   "JR Special Rapid",        "28 min",   "¥580",    "Frequent, Suica-payable, no reservation. Skip the Shinkansen."],
    ["Kyoto ↔ Tokyo",   "Nozomi (Shinkansen)",     "2h 15m",   "¥14,170", "If returning via Kyoto Station."],
    ["Osaka ↔ Nara",    "JR Yamatoji Rapid",       "45 min",   "¥820",    "Direct from Osaka Station (under the hotel)."],
    ["Osaka ↔ USJ",     "JR Yumesaki Line",        "12 min",   "¥190",    "From Osaka Station to Universal City. Direct."],
    ["Narita ↔ Tokyo",  "Narita Express (N'EX)",   "75 min",   "¥3,250",  "Reserved. Runs to Shinagawa / Tokyo Station / Shinjuku."],
    ["Shin-Osaka ↔ Osaka Station", "JR Kyoto Line", "4 min",   "¥170",    "One stop south. Suica-tap. Your Shinkansen connector."],
]

for r_idx, row in enumerate(transit_rows, start=6):
    alt = r_idx % 2 == 0
    for c_idx, val in enumerate(row, start=1):
        cell = s3.cell(row=r_idx, column=c_idx, value=val)
        style_body(cell, alt=alt, bold=(c_idx == 1))

t_widths = [28, 28, 12, 18, 60]
for i, w in enumerate(t_widths, 1):
    s3.column_dimensions[get_column_letter(i)].width = w
for r in range(6, 6 + len(transit_rows)):
    s3.row_dimensions[r].height = 28

# Payments + phrases, stacked below
start = 6 + len(transit_rows) + 2
s3.cell(row=start, column=1, value="Payments").font = Font(name=FONT, size=12, bold=True, color=INK)
s3.merge_cells(start_row=start, start_column=1, end_row=start, end_column=5)

pay_notes = [
    "Add Suica to Apple Wallet before flying — pays for trains, vending, most convenience stores.",
    "Keep ~¥20,000 per person in small bills. Top up at 7-Eleven ATMs (English, accepts foreign cards).",
    "Carry Visa + Mastercard. Amex works at the big hotels, Kikunoi, Den — patchy elsewhere.",
    "Do not tip — anywhere, ever. The price is the price.",
    "Escalators: stand LEFT in Tokyo, RIGHT in Osaka.",
]
for i, line in enumerate(pay_notes, 1):
    c = s3.cell(row=start + i, column=1, value="• " + line)
    c.font = Font(name=FONT, size=10, color=INK_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    s3.merge_cells(start_row=start + i, start_column=1, end_row=start + i, end_column=5)
    s3.row_dimensions[start + i].height = 22

start2 = start + len(pay_notes) + 2
s3.cell(row=start2, column=1, value="Apps before you fly").font = Font(name=FONT, size=12, bold=True, color=INK)
s3.merge_cells(start_row=start2, start_column=1, end_row=start2, end_column=5)

apps = [
    "Google Maps — train navigation, platforms, exits",
    "SmartEX — Shinkansen reservations",
    "GO — Japan's Uber; pays by saved credit card",
    "Tokyo Disney Resort — Premier / Priority passes",
    "Universal Studios Japan — Super Nintendo timed entry",
    "Pocket Concierge — Den and high-end sushi bookings",
    "Google Translate + Papago — camera menu translation",
]
for i, line in enumerate(apps, 1):
    c = s3.cell(row=start2 + i, column=1, value="• " + line)
    c.font = Font(name=FONT, size=10, color=INK_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    s3.merge_cells(start_row=start2 + i, start_column=1, end_row=start2 + i, end_column=5)
    s3.row_dimensions[start2 + i].height = 20

start3 = start2 + len(apps) + 2
s3.cell(row=start3, column=1, value="Phrases worth learning").font = Font(name=FONT, size=12, bold=True, color=INK)
s3.merge_cells(start_row=start3, start_column=1, end_row=start3, end_column=5)

phrases = [
    "Sumimasen — excuse me / sorry / thank you (use 40x a day)",
    "Arigatō gozaimasu — thank you, properly",
    "Onegai shimasu — please, if you would",
    "Oishii! — delicious",
    "Okaikei onegaishimasu — check, please",
    "Eigo no menyū wa arimasu ka? — is there an English menu?",
    "Daijōbu desu — I'm fine / no thank you / all good",
]
for i, line in enumerate(phrases, 1):
    c = s3.cell(row=start3 + i, column=1, value="• " + line)
    c.font = Font(name=FONT, size=10, color=INK_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    s3.merge_cells(start_row=start3 + i, start_column=1, end_row=start3 + i, end_column=5)
    s3.row_dimensions[start3 + i].height = 20

s3.freeze_panes = "A6"
s3.page_setup.orientation = s3.ORIENTATION_LANDSCAPE
s3.page_setup.fitToWidth = 1
s3.page_setup.fitToHeight = 0
s3.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print(f"saved {OUT}")
